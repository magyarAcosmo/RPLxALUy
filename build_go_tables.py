### Gene Ontology Tables
# organize HOMER findGO.pl output into summary tables
# additional manual edits made in Excel (sorting by significance, inclusion of terms, etc) after creating tables here
import os
import re
import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================================
# Paths
# =========================================
BASE        = r"PATH TO DATA FOLDER" # replace with path
WAGE_GO     = os.path.join(BASE, "DMRs_wAGE",  "go_output_allDMRs_wAGE")
NOAGE_GO    = os.path.join(BASE, "DMRs_woAGE", "go_output_allDMRs_woAGE")
WAGE_ANNOT  = os.path.join(BASE, "DMRs_wAGE",  "dmrs_wAGE_genes.txt")
NOAGE_ANNOT = os.path.join(BASE, "DMRs_woAGE", "dmrs_woAGE_genes.txt")
OUTPUT      = os.path.join(BASE, "GO_developmental_summary.xlsx")

PVAL_THRESHOLD = 0.05

# Only biological process for paper focus
GO_FILE = "biological_process.txt"

# Embryonic/developmental filter keywords
EMBRYO_KEYWORDS = [
    "develop",      "embryo",       "morphogen",
    "axon",         "neuron",       "guidance",
    "differentiat", "stem cell",    "pattern",
    "organogen",    "cell fate",    "progenitor",
    "adherens",     "cell adhesion","junction",
    "ephrin",       "cadherin",     "signaling",
    "brain",        "neural",       "cortical",
    "proliferat",   "migration",    "specification",
    "nervous",      "synap",        "projection",
    "tract",        "growth cone",  "axogenesis"
]

# =========================================
# Styles
# =========================================
HEADER_BLUE      = PatternFill("solid", fgColor="1F4E79")
HEADER_GREEN     = PatternFill("solid", fgColor="375623")
HEADER_SHARED    = PatternFill("solid", fgColor="7B2D8B")
HEADER_GENE      = PatternFill("solid", fgColor="833C00")
ROW_BLUE         = PatternFill("solid", fgColor="D6E4F0")
ROW_GREEN        = PatternFill("solid", fgColor="E2EFDA")
ROW_PURPLE       = PatternFill("solid", fgColor="EAD1F5")
ROW_ORANGE       = PatternFill("solid", fgColor="FCE4D6")
ROW_WHITE        = PatternFill("solid", fgColor="FFFFFF")
FONT_HEADER      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_TITLE       = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
FONT_BODY        = Font(name="Calibri", size=10)
FONT_BOLD        = Font(name="Calibri", bold=True, size=10)
FONT_HIGHLIGHT   = Font(name="Calibri", bold=True, size=10, color="7B0000")
ALIGN_CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER      = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin")
)

# =========================================
# Formatting helpers
# =========================================
def autosize_columns(ws, min_width=12, max_width=55):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

def write_title(ws, row, title, num_cols, fill):
    ws.row_dimensions[row].height = 24
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_TITLE
    cell.fill = fill
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row,   end_column=num_cols)

def write_header(ws, row, headers, fill):
    ws.row_dimensions[row].height = 20
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = FONT_HEADER
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

def write_row(ws, row, values, fill, bold=False, highlight_col=None):
    ws.row_dimensions[row].height = 18
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = (FONT_HIGHLIGHT if col == highlight_col
                     else FONT_BOLD if bold else FONT_BODY)
        cell.fill = fill
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER

# =========================================
# Load and filter GO biological process
# =========================================
def load_and_filter_go(go_dir, keywords, pval_thresh):
    """
    Load biological_process.txt, filter to:
    1. p-value < threshold
    2. Term name contains at least one developmental keyword
    Returns sorted dataframe.
    """
    filepath = os.path.join(go_dir, GO_FILE)
    if not os.path.exists(filepath):
        print(f"  Not found: {filepath}")
        return pd.DataFrame()

    df = pd.read_csv(filepath, sep="\t")
    df.columns = [c.strip() for c in df.columns]
    df["Enrichment"] = pd.to_numeric(df["Enrichment"], errors="coerce")

    # Filter by p-value
    df = df[df["Enrichment"] < pval_thresh].copy()

    # Filter by developmental keywords
    pattern = re.compile("|".join(keywords), re.IGNORECASE)
    df = df[df["Term"].apply(lambda t: bool(pattern.search(str(t))))].copy()

    # Sort by p-value
    df = df.sort_values("Enrichment", ascending=True).reset_index(drop=True)

    print(f"  Developmental GO terms after filtering: {len(df)}")
    return df

# =========================================
# Load annotation -> Entrez map
# =========================================
def load_annotation(filepath):
    if not os.path.exists(filepath):
        print(f"  Annotation not found: {filepath}")
        return {}

    df = pd.read_csv(filepath, sep="\t")
    df.columns = [c.strip() for c in df.columns]

    chr_col    = next((c for c in df.columns if c.lower() == "chr"),            None)
    start_col  = next((c for c in df.columns if c.lower() == "start"),          None)
    end_col    = next((c for c in df.columns if c.lower() in ["end", "stop"]),  None)
    entrez_col = next((c for c in df.columns if "entrez" in c.lower()),         None)
    gene_col   = next((c for c in df.columns if "gene name" in c.lower()),      None)
    annot_col  = next((c for c in df.columns if c.lower() == "annotation"),     None)
    dist_col   = next((c for c in df.columns if "distance to tss" in c.lower()),None)
    peak_col   = df.columns[0]

    entrez_map = {}
    for _, row in df.iterrows():
        entrez_raw = str(row[entrez_col]) if entrez_col else ""
        for eid in re.split(r"[;,]", entrez_raw):
            eid = eid.strip()
            if not eid or eid == "nan":
                continue
            info = {
                "PeakID"     : str(row[peak_col])  if peak_col  else "N/A",
                "Chr"        : str(row[chr_col])   if chr_col   else "N/A",
                "Start"      : str(row[start_col]) if start_col else "N/A",
                "End"        : str(row[end_col])   if end_col   else "N/A",
                "Gene Name"  : str(row[gene_col])  if gene_col  else "N/A",
                "Annotation" : str(row[annot_col]) if annot_col else "N/A",
                "Dist TSS"   : str(row[dist_col])  if dist_col  else "N/A",
            }
            entrez_map.setdefault(eid, []).append(info)

    print(f"  Entrez map: {len(entrez_map)} unique genes")
    return entrez_map

# =========================================
# Get DMR info for a gene list
# =========================================
def get_dmr_info(entrez_str, entrez_map):
    regions, annotations = [], []
    for eid in str(entrez_str).split(","):
        eid = eid.strip()
        if eid in entrez_map:
            for d in entrez_map[eid]:
                r = f"{d['Chr']}:{d['Start']}-{d['End']}"
                a = f"{d['Gene Name']} ({d['Annotation']}, {d['Dist TSS']}bp)"
                if r not in regions:    regions.append(r)
                if a not in annotations: annotations.append(a)
    return (
        "; ".join(regions[:5])      if regions      else "N/A",
        "; ".join(annotations[:5])  if annotations  else "N/A"
    )

# =========================================
# Build gene-centric summary
# =========================================
def build_gene_summary(df_wage, df_noage):
    """
    For each gene appearing in developmental GO terms,
    summarize which terms it contributes to in each dataset.
    """
    gene_data = defaultdict(lambda: {
        "wAGE_terms": [], "woAGE_terms": [], "symbol": ""
    })

    for df, key in [(df_wage, "wAGE_terms"), (df_noage, "woAGE_terms")]:
        if df.empty:
            continue
        for _, row in df.iterrows():
            symbols = str(row.get("Gene Symbols", "")).split(",")
            term    = str(row.get("Term", ""))
            for sym in symbols:
                sym = sym.strip()
                if sym:
                    gene_data[sym]["symbol"] = sym
                    if term not in gene_data[sym][key]:
                        gene_data[sym][key].append(term)

    # Build dataframe — only genes appearing in at least one term
    rows = []
    for sym, data in sorted(gene_data.items()):
        wage_terms  = data["wAGE_terms"]
        woage_terms = data["woAGE_terms"]
        in_both     = bool(wage_terms) and bool(woage_terms)
        rows.append({
            "Gene Symbol"              : sym,
            "In DMRs with AGE"         : "YES" if wage_terms  else "NO",
            "In DMRs without AGE"      : "YES" if woage_terms else "NO",
            "Shared Across Datasets"   : "YES" if in_both     else "NO",
            "GO Terms (with AGE)"      : "; ".join(wage_terms[:5]),
            "GO Terms (without AGE)"   : "; ".join(woage_terms[:5]),
            "Total Term Appearances"   : len(wage_terms) + len(woage_terms)
        })

    gene_df = pd.DataFrame(rows)
    if not gene_df.empty:
        gene_df = gene_df.sort_values(
            ["Shared Across Datasets", "Total Term Appearances"],
            ascending=[False, False]
        ).reset_index(drop=True)

    return gene_df

# =========================================
# Write GO summary sheet
# =========================================
def write_go_sheet(ws, df, entrez_map, label, header_fill, row_fill):
    ws.freeze_panes = "A3"
    current_row = 1

    HEADERS = [
        "GO Term ID",
        "Biological Function",
        "P-value",
        "Log P-value",
        "Total Genes in Term",
        "Target Genes Hit",
        "Fraction of Targets",
        "Key Genes",
        "Nearest DMR Regions",
        "DMR Annotations"
    ]

    title = (f"Developmental GO Biological Processes — {label}  |  "
             f"{len(df)} Significant Terms  |  p < {PVAL_THRESHOLD}")
    write_title(ws, current_row, title, len(HEADERS), header_fill)
    current_row += 1
    write_header(ws, current_row, HEADERS, header_fill)
    current_row += 1

    if df.empty:
        write_row(ws, current_row,
                  ["—"] * 2 + ["No significant developmental GO terms found"] + ["—"] * 8,
                  ROW_WHITE)
        return

    for i, (_, row) in enumerate(df.iterrows()):
        fill   = row_fill if i % 2 == 0 else ROW_WHITE
        pval   = row.get("Enrichment", "N/A")
        try:
            pval_fmt = f"{float(pval):.2e}"
        except:
            pval_fmt = str(pval)

        entrez_str   = str(row.get("Entrez Gene IDs", ""))
        gene_symbols = str(row.get("Gene Symbols",    "N/A"))
        dmr_regions, dmr_annots = get_dmr_info(entrez_str, entrez_map)

        values = [
            str(row.get("TermID",  "N/A")),
            str(row.get("Term",    "N/A")),
            pval_fmt,
            str(row.get("logP",    "N/A")),
            str(row.get("Genes in Term",        "N/A")),
            str(row.get("Target Genes in Term", "N/A")),
            str(row.get("Fraction of Targets in Term", "N/A")),
            gene_symbols,
            dmr_regions,
            dmr_annots
        ]
        write_row(ws, current_row, values, fill)
        current_row += 1

    autosize_columns(ws)
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["I"].width = 40
    ws.column_dimensions["J"].width = 45
    ws.column_dimensions["K"].width = 55

# =========================================
# Main
# =========================================
print("Loading data...")

df_wage  = load_and_filter_go(WAGE_GO,  EMBRYO_KEYWORDS, PVAL_THRESHOLD)
df_noage = load_and_filter_go(NOAGE_GO, EMBRYO_KEYWORDS, PVAL_THRESHOLD)

annot_wage  = load_annotation(WAGE_ANNOT)
annot_noage = load_annotation(NOAGE_ANNOT)

print(f"\nDevelopmental terms — with AGE:    {len(df_wage)}")
print(f"Developmental terms — without AGE: {len(df_noage)}")

# Build gene summary
gene_summary = build_gene_summary(df_wage, df_noage)
shared_genes = gene_summary[gene_summary["Shared Across Datasets"] == "YES"]
print(f"Genes shared across both datasets: {len(shared_genes)}")

# =========================================
# Build workbook
# =========================================
wb = Workbook()
wb.remove(wb.active)

# Sheet 1 — wAGE GO terms
ws1 = wb.create_sheet("GO — DMRs with AGE")
write_go_sheet(ws1, df_wage, annot_wage,
               "DMRs with AGE", HEADER_BLUE, ROW_BLUE)

# Sheet 2 — woAGE GO terms
ws2 = wb.create_sheet("GO — DMRs without AGE")
write_go_sheet(ws2, df_noage, annot_noage,
               "DMRs without AGE", HEADER_GREEN, ROW_GREEN)

# Sheet 3 — Gene-centric summary
ws3 = wb.create_sheet("Key Genes Summary")
ws3.freeze_panes = "A3"
current_row = 1

GENE_HEADERS = [
    "Gene Symbol",
    "In DMRs with AGE", "In DMRs without AGE", "Shared Across Datasets",
    "GO Terms (with AGE)", "GO Terms (without AGE)", "Total Term Appearances"
]

write_title(ws3, current_row,
            f"Key Developmental Genes — Across Both DMR Datasets  |  "
            f"{len(shared_genes)} Shared Genes  |  "
            f"{len(gene_summary)} Total",
            len(GENE_HEADERS), HEADER_SHARED)
current_row += 1
write_header(ws3, current_row, GENE_HEADERS, HEADER_SHARED)
current_row += 1

for i, (_, row) in enumerate(gene_summary.iterrows()):
    is_shared = row["Shared Across Datasets"] == "YES"
    fill      = ROW_PURPLE if is_shared and i % 2 == 0 else (
                ROW_WHITE  if not is_shared else ROW_PURPLE)

    values = [
        row["Gene Symbol"],
        row["In DMRs with AGE"],
        row["In DMRs without AGE"],
        row["Shared Across Datasets"],
        row["GO Terms (with AGE)"],
        row["GO Terms (without AGE)"],
        row["Total Term Appearances"]
    ]
    write_row(ws3, current_row, values, fill,
              highlight_col=5 if is_shared else None)
    current_row += 1

autosize_columns(ws3)
ws3.column_dimensions["F"].width = 55
ws3.column_dimensions["G"].width = 55

wb.save(OUTPUT)
print(f"\nSaved: {OUTPUT}")